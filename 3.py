from openpyxl import *
import os
import psutil
import function
from tkinter import messagebox

while True:  # ввод размерности матрицы
    n = input('Введите размерность квадратной матрицы: ')
    if n.isdigit() and n != '0' and n != '1':
        n = int(n)
        break
    else:
        print('Число введено не корректно!')
i, fg = 1, False
directory = os.getcwd()
matr = []  # исходная матрица
name_file = input('Введите название файла: ')
while not fg:
    num = matr
    function.close_excel()  # закрытие программы, если открыта
    book = Workbook()  # создание шаблона
    book.remove(book['Sheet'])
    worksheet = book.create_sheet('Ввод данных')
    worksheet.append([x if x != 0 else '' for x in range(n+1)])
    for j in range(n):
        mas = [j+1]
        mas.extend([y for i in range(len(num)) for y in num[j] if i == j and len(num) != 0])
        worksheet.append(mas)
    function.close_excel()
    name_file = function.saving_excel(book, name_file, f'({i})')
    book.close()
    os.startfile(directory + '/' + name_file + f'({i}).xlsx')  # ввод данных пользователем
    while 'EXCEL.EXE' in (m.name() for m in psutil.process_iter()):
        pass
    book = load_workbook(name_file + f'({i}).xlsx')
    function.removing_sheets(book, 'Ввод данных')
    num = function.reading_data(book, n)
    num, fg = function.data_checking(num)
    i += 1
    if not fg:  # поиск ошибок и вывод о них
        error = set(j + 1 for j in range(len(num)) for y in num[j] if y is None)
        if len(error) > 1:
            messagebox.showerror(title='Ошибка заполнения', message=f'Ошибка в следующих строках: {str(error)[1:-1]}')
        else:
            messagebox.showerror(title='Ошибка заполнения', message=f'Ошибка в следующей строке: {str(error)[1:-1]}')
    matr = num
matrix = []  # создание преобразованной матрицы
for i in range(len(matr)):
    line = []
    for j in range(len(matr)):
        line.append(function.transformation(matr[i][j], i, j, matr))
    matrix.append(line)
book = Workbook()
book.remove(book['Sheet'])
function.data_output(book, 'Ввод данных', 'Исходная матрица', matr, n)
function.data_output(book, 'Вывод данных', 'Преобразованная матрица', matrix, n)
name_file = function.saving_excel(book, '3.xlsx', '')
os.startfile(directory + '/3.xlsx')
