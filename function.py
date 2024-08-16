from openpyxl import *
import os
import psutil
from tkinter import messagebox





def close_excel():
    if 'EXCEL.EXE' in (m.name() for m in psutil.process_iter()):
        command = 'taskkill /IM Excel.exe'
        os.system(command)


def deleting_files(name_file, i):
    while os.path.isfile(name_file + f'({i - 1}).xlsx'):
        try:
            os.remove(name_file + f'({i - 1}).xlsx')
        except PermissionError:
            command = 'taskkill /IM Excel.exe'
            os.system(command)


def saving_excel(file, txt, i):
    while True:
        try:
            if i != '':
                file.save(txt + i + '.xlsx')
            else:
                file.save(txt)
            return txt
        except PermissionError:
            messagebox.showerror('Ошибка', 'Файл не может быть создан')
            exit()
        except:
            messagebox.showerror('Ошибка', 'Файл не может так называться')
            messagebox.showinfo('Сообщение', 'Файлу будет присовено другое имя')
            txt = '31'


def removing_sheets(file, name_sheet):
    sheets = file.sheetnames
    for sheet in sheets:
        if sheet != name_sheet:
            file.remove(file[name_sheet])


def reading_data(file, n):
    worksheet = file.active
    massive = []
    for row in range(2, n + 2):
        line = []
        for column in range(2, n + 2):
            cell = worksheet.cell(row=row, column=column).value
            line.append(cell)
        massive.append(line)
    return massive


def data_checking(massive):
    matr = []
    flag = True
    for i in range(len(massive)):
        row = []
        for j in range(len(massive)):
            try:
                q = float(massive[i][j])
                row.append(q)
            except (ValueError, TypeError):
                row.append(None)
                flag = False
        matr.append(row)
    return matr, flag


def transformation(q, index1, index2, massive):
    quantity, amount = 0, 0
    for i in range(index1-1, index1+2):
        for j in range(index2-1, index2+2):
            if i >= 0 and j >= 0:
                try:
                    amount += massive[i][j]
                    quantity += 1
                except IndexError:
                    pass
    return (amount - q) / (quantity - 1)


def data_output(file, txt1, txt2, massive, n):
    worksheet = file.create_sheet(txt1)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n + 1)
    worksheet['A1'] = txt2
    worksheet.append([x if x != 0 else 'n' for x in range(n + 1)])
    for j in range(len(massive)):
        line = [j + 1]
        for y in massive[j]:
            line.append(y)
        worksheet.append(line)
